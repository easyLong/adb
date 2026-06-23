"""Read-only KOL daily metrics web page."""

from __future__ import annotations

import argparse
import html
import io
import json
from datetime import date, datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse

from apps.finance_crawler.crawler_app.storage.db import get_conn, init_crawler_app_db


DEFAULT_HOST = "0.0.0.0"
DEFAULT_PORT = 8091
MAX_LIMIT = 2000
SORT_OPTIONS = {
    "base_id": "CASE WHEN b.id IS NULL THEN 1 ELSE 0 END, b.id ASC, m.metric_date DESC, m.platform ASC, m.kol_name ASC",
    "title": "m.kol_name ASC, m.metric_date DESC",
    "title_desc": "m.kol_name DESC, m.metric_date DESC",
    "date_desc": "m.metric_date DESC, m.platform ASC, b.group_name ASC, m.kol_name ASC",
    "date_asc": "m.metric_date ASC, m.platform ASC, b.group_name ASC, m.kol_name ASC",
    "platform": "m.platform ASC, b.group_name ASC, m.kol_name ASC, m.metric_date DESC",
    "group": "b.group_name ASC, m.kol_name ASC, m.metric_date DESC",
    "fans_desc": "m.fans_count DESC, m.metric_date DESC, m.kol_name ASC",
    "growth_desc": "m.growth_count DESC, m.metric_date DESC, m.kol_name ASC",
    "read_desc": "m.read_count DESC, m.metric_date DESC, m.kol_name ASC",
}
SORT_LABELS = {
    "base_id": "基础表顺序",
    "title": "Title 升序",
    "title_desc": "Title 降序",
    "date_desc": "日期最新",
    "date_asc": "日期最早",
    "platform": "平台",
    "group": "群",
    "fans_desc": "粉丝数高到低",
    "growth_desc": "增粉数高到低",
    "read_desc": "阅读数高到低",
}
MISSING_OPTIONS = {
    "": "全部",
    "fans_empty": "粉丝数为空",
    "growth_empty": "增粉数为空",
    "fans_or_growth_empty": "粉丝数或增粉数为空",
    "fans_and_growth_empty": "粉丝数和增粉数都为空",
}


def main() -> int:
    parser = argparse.ArgumentParser(description="KOL daily metrics web page")
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    args = parser.parse_args()

    init_crawler_app_db()
    server = ThreadingHTTPServer((args.host, args.port), KolMetricsHandler)
    print(f"KOL metrics web: http://{args.host}:{args.port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("KOL metrics web stopped")
    finally:
        server.server_close()
    return 0


class KolMetricsHandler(BaseHTTPRequestHandler):
    server_version = "KolMetricsWeb/1.0"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path in {"", "/"}:
            self._send_html(render_page(_params(parsed.query)))
            return
        if parsed.path == "/api/metrics":
            self._send_json(load_page_data(_params(parsed.query)))
            return
        if parsed.path == "/download.xlsx":
            self._send_xlsx(export_xlsx(_params(parsed.query)))
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def log_message(self, format: str, *args: Any) -> None:
        print("%s - %s" % (self.address_string(), format % args))

    def _send_html(self, body: str) -> None:
        data = body.encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _send_json(self, value: Any) -> None:
        data = json.dumps(value, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _send_xlsx(self, data: bytes) -> None:
        today = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.send_response(HTTPStatus.OK)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Disposition", f'attachment; filename="kol_daily_metrics_{today}.xlsx"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def _params(query: str) -> dict[str, str]:
    raw = parse_qs(query, keep_blank_values=True)
    return {key: values[-1].strip() for key, values in raw.items()}


def load_page_data(params: dict[str, str]) -> dict[str, Any]:
    filters = _normalize_filters(params)
    conn = get_conn()
    try:
        return {
            "filters": filters,
            "options": _load_options(conn),
            "summary": _load_summary(conn, filters),
            "rows": _load_rows(conn, filters),
        }
    finally:
        conn.close()


def export_xlsx(params: dict[str, str]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Missing dependency: pip install openpyxl") from exc

    filters = _normalize_filters(params)
    conn = get_conn()
    try:
        rows = _load_rows(conn, filters)
    finally:
        conn.close()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "kol_daily_metrics"
    columns = export_columns()
    worksheet.append([title for title, _key in columns])
    for row in rows:
        worksheet.append([excel_value(row.get(key)) for _title, key in columns])

    header_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_cells in worksheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="center")

    widths = [12, 24, 10, 10, 10, 45, 12, 12, 12, 12, 20, 18, 36]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _normalize_filters(params: dict[str, str]) -> dict[str, Any]:
    limit = _parse_int(params.get("limit"), 500) or 500
    sort = params.get("sort", "base_id")
    if sort not in SORT_OPTIONS:
        sort = "base_id"
    return {
        "metric_date": _parse_date(params.get("date")),
        "platform": params.get("platform", ""),
        "kol_type": params.get("kol_type", ""),
        "missing": params.get("missing", "") if params.get("missing", "") in MISSING_OPTIONS else "",
        "q": params.get("q", ""),
        "limit": min(max(limit, 1), MAX_LIMIT),
        "sort": sort,
    }


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_int(value: str | None, default: int | None = None) -> int | None:
    if value in {None, ""}:
        return default
    try:
        return int(str(value))
    except ValueError:
        return default


def _load_options(conn) -> dict[str, Any]:
    with conn.cursor() as cursor:
        cursor.execute(
            """
            SELECT DISTINCT metric_date
            FROM kol_daily_metrics
            ORDER BY metric_date DESC
            """
        )
        dates = [row["metric_date"].isoformat() for row in cursor.fetchall()]
        cursor.execute(
            """
            SELECT DISTINCT platform
            FROM kol_daily_metrics
            ORDER BY platform
            """
        )
        platforms = [str(row["platform"]) for row in cursor.fetchall()]
        cursor.execute(
            """
            SELECT DISTINCT COALESCE(NULLIF(b.kol_type, ''), '未匹配') AS kol_type
            FROM kol_daily_metrics m
            LEFT JOIN kol_base_profiles b
              ON b.kol_name = m.kol_name
             AND b.platform = m.platform
            ORDER BY kol_type
            """
        )
        kol_types = [str(row["kol_type"]) for row in cursor.fetchall()]
    return {
        "dates": dates,
        "platforms": platforms,
        "kol_types": kol_types,
    }


def _where_clause(filters: dict[str, Any]) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    args: list[Any] = []
    if filters["metric_date"]:
        clauses.append("m.metric_date = %s")
        args.append(filters["metric_date"])
    if filters["platform"]:
        clauses.append("m.platform = %s")
        args.append(filters["platform"])
    if filters["kol_type"]:
        clauses.append("COALESCE(NULLIF(b.kol_type, ''), '未匹配') = %s")
        args.append(filters["kol_type"])
    if filters["missing"] == "fans_empty":
        clauses.append("m.fans_count IS NULL")
    elif filters["missing"] == "growth_empty":
        clauses.append("m.growth_count IS NULL")
    elif filters["missing"] == "fans_or_growth_empty":
        clauses.append("(m.fans_count IS NULL OR m.growth_count IS NULL)")
    elif filters["missing"] == "fans_and_growth_empty":
        clauses.append("(m.fans_count IS NULL AND m.growth_count IS NULL)")
    if filters["q"]:
        clauses.append("(m.kol_name LIKE %s OR b.group_name LIKE %s)")
        like = f"%{filters['q']}%"
        args.extend([like, like])
    return ("WHERE " + " AND ".join(clauses), args) if clauses else ("", args)


def _load_summary(conn, filters: dict[str, Any]) -> dict[str, Any]:
    where_sql, args = _where_clause(filters)
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                COUNT(*) AS total_rows,
                COUNT(DISTINCT m.metric_date) AS date_count,
                COUNT(DISTINCT m.kol_name) AS kol_count,
                SUM(m.fans_count IS NOT NULL) AS fans_rows,
                SUM(m.growth_count IS NOT NULL) AS growth_rows,
                SUM(m.read_count IS NOT NULL) AS read_rows,
                SUM(m.post_count_24h IS NOT NULL) AS post_rows,
                SUM(COALESCE(NULLIF(b.kol_type, ''), '未匹配') = '内部') AS internal_rows,
                SUM(b.id IS NULL) AS unmatched_base_rows
            FROM kol_daily_metrics m
            LEFT JOIN kol_base_profiles b
              ON b.kol_name = m.kol_name
             AND b.platform = m.platform
            {where_sql}
            """,
            args,
        )
        row = cursor.fetchone() or {}
    return {key: int(value or 0) for key, value in row.items()}


def _load_rows(conn, filters: dict[str, Any]) -> list[dict[str, Any]]:
    where_sql, args = _where_clause(filters)
    order_sql = SORT_OPTIONS.get(filters["sort"], SORT_OPTIONS["title"])
    args = [*args, int(filters["limit"])]
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                m.metric_date,
                m.kol_name,
                m.platform,
                b.homepage_url,
                b.group_name,
                COALESCE(NULLIF(b.kol_type, ''), '未匹配') AS kol_type,
                m.fans_count,
                m.growth_count,
                m.read_count,
                m.post_count_24h,
                m.source_payload_json,
                m.writeback_error,
                m.updated_at
            FROM kol_daily_metrics m
            LEFT JOIN kol_base_profiles b
              ON b.kol_name = m.kol_name
             AND b.platform = m.platform
            {where_sql}
            ORDER BY {order_sql}
            LIMIT %s
            """,
            args,
        )
        rows = [dict(row) for row in cursor.fetchall()]
        for row in rows:
            row["remark"] = row_remark(row)
        return rows


def render_page(params: dict[str, str]) -> str:
    data = load_page_data(params)
    filters = data["filters"]
    options = data["options"]
    summary = data["summary"]
    rows = data["rows"]
    date_value = filters["metric_date"].isoformat() if filters["metric_date"] else ""

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>KOL Daily Metrics</title>
  <style>
    :root {{
      color-scheme: light;
      --text: #1f2933;
      --muted: #637083;
      --line: #d9dee7;
      --head: #eef2f7;
      --cell: #cfd6df;
      --accent: #1769aa;
      --ok: #11693a;
      --bad: #a33b2f;
      --warn: #8a6200;
      --bg: #f3f5f8;
      --panel: #ffffff;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font: 14px/1.45 "Segoe UI", "Microsoft YaHei", Arial, sans-serif;
      color: var(--text);
      background: var(--bg);
    }}
    header {{
      background: var(--panel);
      border-bottom: 1px solid var(--line);
      padding: 12px 16px;
    }}
    h1 {{
      margin: 0 0 8px;
      font-size: 18px;
      font-weight: 650;
      letter-spacing: 0;
    }}
    form {{
      display: grid;
      grid-template-columns: repeat(7, minmax(110px, 1fr)) auto;
      gap: 8px;
      align-items: end;
    }}
    label {{
      display: grid;
      gap: 4px;
      color: var(--muted);
      font-size: 12px;
    }}
    select, input, button {{
      height: 30px;
      border: 1px solid var(--line);
      border-radius: 3px;
      background: #fff;
      color: var(--text);
      padding: 0 8px;
      font: inherit;
      min-width: 0;
    }}
    button {{
      background: var(--accent);
      border-color: var(--accent);
      color: #fff;
      cursor: pointer;
      padding: 0 14px;
    }}
    .actions {{
      display: flex;
      gap: 8px;
      align-items: end;
      min-width: 160px;
    }}
    .download {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      height: 30px;
      padding: 0 12px;
      border: 1px solid var(--line);
      border-radius: 3px;
      background: #fff;
      color: var(--accent);
      white-space: nowrap;
    }}
    .download.copied {{
      border-color: var(--ok);
      color: var(--ok);
    }}
    main {{ padding: 12px 16px 24px; }}
    .result-meta {{
      background: var(--panel);
      border: 1px solid var(--cell);
      border-bottom: 0;
      padding: 7px 9px;
      color: var(--muted);
      font-size: 13px;
    }}
    .table-wrap {{
      overflow: auto;
      border: 1px solid var(--cell);
      background: var(--panel);
    }}
    table {{
      border-collapse: collapse;
      width: 100%;
      min-width: 1120px;
      font-size: 13px;
    }}
    th, td {{
      border-right: 1px solid var(--cell);
      border-bottom: 1px solid var(--cell);
      padding: 5px 8px;
      text-align: left;
      white-space: nowrap;
      vertical-align: middle;
    }}
    th:last-child, td:last-child {{ border-right: 0; }}
    tbody tr:last-child td {{ border-bottom: 0; }}
    th {{
      background: var(--head);
      font-size: 12px;
      color: #3f4d5d;
      font-weight: 650;
    }}
    tbody tr:nth-child(even) {{ background: #fbfcfe; }}
    tbody tr:hover {{ background: #eaf3ff; }}
    .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
    .muted {{ color: var(--muted); }}
    .status-success {{ color: var(--ok); font-weight: 600; }}
    .status-error {{ color: var(--bad); font-weight: 600; }}
    .status-pending, .status-synced-from-doc {{ color: var(--warn); font-weight: 600; }}
    a {{ color: var(--accent); text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    @media (max-width: 900px) {{
      header {{ padding: 12px; }}
      main {{ padding: 12px; }}
      form {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>KOL Daily Metrics</h1>
    {render_filters(date_value, filters, options)}
  </header>
  <main>
    {render_result_meta(summary, len(rows), filters)}
    <div class="table-wrap">
      <table>
        <thead>{render_table_header()}</thead>
        <tbody>{''.join(render_row(row) for row in rows) or '<tr><td colspan="13" class="muted">没有数据</td></tr>'}</tbody>
      </table>
    </div>
  </main>
  {render_copy_script(filters, date_value)}
</body>
</html>"""


def render_filters(date_value: str, filters: dict[str, Any], options: dict[str, Any]) -> str:
    return f"""<form method="get">
  <label>日期{select("date", date_value, options["dates"], "全部日期")}</label>
  <label>平台{select("platform", filters["platform"], options["platforms"], "全部平台")}</label>
  <label>类型{select("kol_type", filters["kol_type"], options["kol_types"], "全部类型")}</label>
  <label>空值筛选{select("missing", filters["missing"], list(MISSING_OPTIONS), "")}</label>
  <label>排序{select("sort", filters["sort"], list(SORT_LABELS), "")}</label>
  <label>搜索<input name="q" value="{e(filters['q'])}" placeholder="大V名称 / 群"></label>
  <label>行数<input name="limit" value="{int(filters['limit'])}" inputmode="numeric"></label>
  <div class="actions"><button type="submit">查询</button><button type="button" class="download" id="copy-table">复制表格</button><a class="download" href="{download_href(filters, date_value)}">下载 Excel</a></div>
</form>"""


def select(name: str, value: str, options: list[str], blank: str) -> str:
    items = [f'<option value="">{e(blank)}</option>'] if blank else []
    for option in options:
        selected = " selected" if str(option) == str(value) else ""
        label = SORT_LABELS.get(str(option), MISSING_OPTIONS.get(str(option), str(option)))
        items.append(f'<option value="{e(option)}"{selected}>{e(label)}</option>')
    return f'<select name="{e(name)}">{"".join(items)}</select>'


def render_result_meta(summary: dict[str, int], shown_rows: int, filters: dict[str, Any]) -> str:
    parts = [
        f"显示 {shown_rows} 行",
        f"总计 {summary.get('total_rows', 0)} 行",
        f"{summary.get('date_count', 0)} 个日期",
        f"{summary.get('kol_count', 0)} 个大V",
        f"粉丝数 {summary.get('fans_rows', 0)} 行",
        f"阅读数 {summary.get('read_rows', 0)} 行",
        f"未匹配基础 {summary.get('unmatched_base_rows', 0)} 行",
        f"空值筛选：{MISSING_OPTIONS.get(str(filters.get('missing')), '全部')}",
        f"排序：{SORT_LABELS.get(str(filters.get('sort')), str(filters.get('sort') or ''))}",
    ]
    return f'<div class="result-meta">{" | ".join(e(part) for part in parts)}</div>'


def render_table_header() -> str:
    headers = [title for title, _key in export_columns()]
    return "<tr>" + "".join(f"<th>{e(item)}</th>" for item in headers) + "</tr>"


def export_columns() -> list[tuple[str, str]]:
    return [
        ("日期", "metric_date"),
        ("Title/大V名称", "kol_name"),
        ("平台", "platform"),
        ("类型", "kol_type"),
        ("第几群", "group_name"),
        ("主页", "homepage_url"),
        ("粉丝数", "fans_count"),
        ("增粉数", "growth_count"),
        ("阅读数", "read_count"),
        ("24h发文", "post_count_24h"),
        ("更新时间", "updated_at"),
        ("备注", "remark"),
        ("错误", "writeback_error"),
    ]


def render_row(row: dict[str, Any]) -> str:
    homepage = str(row.get("homepage_url") or "")
    homepage_cell = f'<a href="{e(homepage)}" target="_blank" rel="noreferrer">打开</a>' if homepage else ""
    cells = [
        row.get("metric_date"),
        row.get("kol_name"),
        row.get("platform"),
        row.get("kol_type"),
        row.get("group_name"),
        homepage_cell,
        n(row.get("fans_count")),
        n(row.get("growth_count")),
        n(row.get("read_count")),
        n(row.get("post_count_24h")),
        row.get("updated_at"),
        row.get("remark"),
        row.get("writeback_error"),
    ]
    numeric_indexes = {6, 7, 8, 9}
    rendered = []
    for index, value in enumerate(cells):
        css = ' class="num"' if index in numeric_indexes else ""
        if isinstance(value, str) and value.startswith("<a "):
            rendered.append(f"<td{css}>{value}</td>")
        elif isinstance(value, str) and value.startswith("<span "):
            rendered.append(f"<td{css}>{value}</td>")
        else:
            rendered.append(f"<td{css}>{e(value)}</td>")
    return "<tr>" + "".join(rendered) + "</tr>"


def row_remark(row: dict[str, Any]) -> str:
    payload = json_loads(row.get("source_payload_json")) or {}
    warning = str(payload.get("quality_warning") or "").strip()
    if warning:
        return warning
    if payload.get("nickname_mismatch"):
        return "昵称不一致"
    return ""


def render_copy_script(filters: dict[str, Any], date_value: str) -> str:
    config = {
        "apiUrl": api_href(filters, date_value),
        "columns": export_columns(),
    }
    config_json = json.dumps(config, ensure_ascii=False).replace("</", "<\\/")
    script = r"""<script>
(function () {
  const config = __COPY_CONFIG__;
  const button = document.getElementById("copy-table");
  if (!button) {
    return;
  }

  function cell(value) {
    if (value === null || value === undefined) {
      return "";
    }
    return String(value).replace(/\r?\n/g, " ").replace(/\t/g, " ").trim();
  }

  function rowsToTsv(rows) {
    const lines = [config.columns.map(function (column) { return column[0]; }).join("\t")];
    rows.forEach(function (row) {
      lines.push(config.columns.map(function (column) { return cell(row[column[1]]); }).join("\t"));
    });
    return lines.join("\r\n");
  }

  async function writeClipboard(text) {
    if (navigator.clipboard && window.isSecureContext) {
      await navigator.clipboard.writeText(text);
      return;
    }
    const textarea = document.createElement("textarea");
    textarea.value = text;
    textarea.setAttribute("readonly", "");
    textarea.style.position = "fixed";
    textarea.style.left = "-9999px";
    document.body.appendChild(textarea);
    textarea.select();
    document.execCommand("copy");
    textarea.remove();
  }

  function setButtonText(text, copied) {
    button.textContent = text;
    button.classList.toggle("copied", Boolean(copied));
  }

  button.addEventListener("click", async function () {
    const original = button.textContent;
    button.disabled = true;
    setButtonText("复制中...", false);
    try {
      const response = await fetch(config.apiUrl, { credentials: "same-origin" });
      if (!response.ok) {
        throw new Error("HTTP " + response.status);
      }
      const data = await response.json();
      const rows = Array.isArray(data.rows) ? data.rows : [];
      await writeClipboard(rowsToTsv(rows));
      setButtonText("已复制", true);
      setTimeout(function () { setButtonText(original || "复制表格", false); }, 1600);
    } catch (error) {
      console.error(error);
      setButtonText("复制失败", false);
      setTimeout(function () { setButtonText(original || "复制表格", false); }, 2200);
    } finally {
      button.disabled = false;
    }
  });
})();
</script>"""
    return script.replace("__COPY_CONFIG__", config_json)


def api_href(filters: dict[str, Any], date_value: str) -> str:
    query = {
        "date": date_value,
        "platform": filters["platform"],
        "kol_type": filters["kol_type"],
        "missing": filters["missing"],
        "sort": filters["sort"],
        "q": filters["q"],
        "limit": int(filters["limit"]),
    }
    return "/api/metrics?" + urlencode({key: value for key, value in query.items() if value not in {None, ""}})


def download_href(filters: dict[str, Any], date_value: str) -> str:
    query = {
        "date": date_value,
        "platform": filters["platform"],
        "kol_type": filters["kol_type"],
        "missing": filters["missing"],
        "sort": filters["sort"],
        "q": filters["q"],
        "limit": int(filters["limit"]),
    }
    return "/download.xlsx?" + urlencode({key: value for key, value in query.items() if value not in {None, ""}})


def excel_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value
    return "" if value is None else value


def json_loads(value: Any) -> Any:
    if not value:
        return None
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return None


def n(value: Any) -> str:
    if value is None:
        return ""
    try:
        return f"{int(value):,}"
    except (TypeError, ValueError):
        return str(value)


def e(value: Any) -> str:
    if value is None:
        return ""
    return html.escape(str(value), quote=True)


if __name__ == "__main__":
    raise SystemExit(main())
