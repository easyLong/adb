"""Direct detail-crawl workflow for local Excel workbooks.

This mode does not run a separate initial check. It reads links from Excel,
creates task-center submissions/executions for each row, crawls each link once,
and writes results back to an output workbook.
"""

from __future__ import annotations

import json
import shutil
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from apps.finance_crawler.config import Config
from apps.finance_crawler.crawlers.constants import (
    SOURCE_ALIPAY,
    SOURCE_ANTFORTUNE,
    SOURCE_TENPAY,
    SOURCE_UNKNOWN,
)
from apps.finance_crawler.mobile.crawler import (
    open_url,
    reset_device_session,
    resolve_short_url,
    scrape_record_content,
)
from apps.finance_crawler.domain.task_types import DETAIL_CRAWL_TASK_TYPE
from apps.finance_crawler.storage.framework_db import (
    finish_task_execution,
    start_task_execution,
    update_task_execution_writeback,
    upsert_excel_row_submission,
)
from apps.finance_crawler.utils.device_health import DeviceUnavailable, assert_device_ready
from apps.finance_crawler.utils.link_source import detect_link_source
from apps.finance_crawler.utils.logger import get_logger

logger = get_logger("excel_detail_workflow")

_SOURCE_LIMITS = {
    SOURCE_ALIPAY: Config.EXCEL_DETAIL_ALIPAY_LIMIT,
    SOURCE_ANTFORTUNE: Config.EXCEL_DETAIL_ANTFORTUNE_LIMIT,
    SOURCE_TENPAY: Config.EXCEL_DETAIL_TENPAY_LIMIT,
}


@dataclass(frozen=True)
class ExcelDetailTarget:
    row_index: int
    source_app: str
    url: str


def run_local_excel_detail() -> list[dict[str, Any]]:
    """Run direct detail crawling for a configured local Excel workbook."""

    start_time = time.time()
    input_path = _input_path()
    output_path = _output_path(input_path)
    result_jsonl_path = _result_jsonl_path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    result_jsonl_path.parent.mkdir(parents=True, exist_ok=True)
    if input_path.resolve() != output_path.resolve():
        shutil.copy2(input_path, output_path)
    if result_jsonl_path.exists():
        result_jsonl_path.unlink()

    workbook, worksheet = _open_workbook(output_path)
    results: list[dict[str, Any]] = []
    try:
        _ensure_result_headers(worksheet)
        targets = _pick_targets(worksheet)
        logger.info(
            "Excel detail started file=%s sheet=%s targets=%s",
            output_path,
            worksheet.title,
            len(targets),
        )
        print(
            json.dumps(
                {"targets": len(targets), "output": str(output_path), "results": str(result_jsonl_path)},
                ensure_ascii=True,
            )
        )

        if not targets:
            return []

        assert_device_ready()

        buffered_writebacks: list[tuple[int, dict[str, Any]]] = []
        for index, target in enumerate(targets, start=1):
            execution_id = _start_execution_for_target(input_path, output_path, worksheet.title, target)
            if execution_id is None:
                logger.info("Excel detail skipped by task submission state row=%s", target.row_index)
                continue
            try:
                item = _crawl_target(index, len(targets), target)
            except DeviceUnavailable as exc:
                _finish_excel_execution(
                    execution_id,
                    item=_error_item(target, str(exc)),
                    writeback_status="skipped",
                    writeback_locator={"path": str(output_path.resolve()), "row_index": target.row_index},
                )
                raise
            results.append(item)
            _write_result(worksheet, item)
            _append_jsonl(result_jsonl_path, item)
            writeback_locator = {"path": str(output_path.resolve()), "row_index": target.row_index}
            if _should_save(index):
                workbook.save(output_path)
                writeback_status = "success"
            else:
                writeback_status = "buffered"
                buffered_writebacks.append((execution_id, writeback_locator))
            _finish_excel_execution(
                execution_id,
                item=item,
                writeback_status=writeback_status,
                writeback_locator=writeback_locator,
            )
            print(json.dumps(item, ensure_ascii=True), flush=True)

        workbook.save(output_path)
        for execution_id, writeback_locator in buffered_writebacks:
            _update_excel_writeback(
                execution_id,
                writeback_status="success",
                writeback_locator=writeback_locator,
            )
        summary = _summary(results, time.time() - start_time)
        _append_jsonl(result_jsonl_path, {"summary": summary})
        logger.info("Excel detail finished: %s", summary)
        print(json.dumps({"summary": summary}, ensure_ascii=True))
        return results
    finally:
        workbook.close()


def _input_path() -> Path:
    if not Config.EXCEL_DETAIL_INPUT_PATH:
        raise ValueError("EXCEL_DETAIL_INPUT_PATH is required for excel-detail")
    path = Path(Config.EXCEL_DETAIL_INPUT_PATH)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    return path


def _output_path(input_path: Path) -> Path:
    if Config.EXCEL_DETAIL_OUTPUT_PATH:
        return Path(Config.EXCEL_DETAIL_OUTPUT_PATH)
    return input_path.with_name(f"{input_path.stem}_detail_output{input_path.suffix}")


def _result_jsonl_path(output_path: Path) -> Path:
    if Config.EXCEL_DETAIL_RESULT_JSONL_PATH:
        return Path(Config.EXCEL_DETAIL_RESULT_JSONL_PATH)
    return output_path.with_name(f"{output_path.stem}_results.jsonl")


def _open_workbook(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Missing dependency: pip install openpyxl") from exc

    workbook = load_workbook(path)
    worksheet = workbook[Config.EXCEL_DETAIL_SHEET_NAME] if Config.EXCEL_DETAIL_SHEET_NAME else workbook.active
    return workbook, worksheet


def _pick_targets(worksheet) -> list[ExcelDetailTarget]:
    targets: list[ExcelDetailTarget] = []
    source_counts: Counter[str] = Counter()
    source_filter = _source_filter()
    for row_index in range(2, worksheet.max_row + 1):
        url = _cell_value(worksheet, row_index, Config.EXCEL_DETAIL_COL_URL)
        if not url:
            continue
        source_app = detect_link_source(url)
        if source_app == SOURCE_UNKNOWN:
            logger.warning("Excel row skipped: unsupported source row=%s url=%s", row_index, url)
            continue
        if source_filter and source_app not in source_filter:
            continue
        if Config.EXCEL_DETAIL_ONLY_EMPTY and _has_existing_result(worksheet, row_index):
            continue
        if _source_limit_reached(source_app, source_counts[source_app]):
            continue
        if Config.EXCEL_DETAIL_LIMIT and len(targets) >= Config.EXCEL_DETAIL_LIMIT:
            break
        source_counts[source_app] += 1
        targets.append(ExcelDetailTarget(row_index=row_index, source_app=source_app, url=url))
    return targets


def _crawl_target(index: int, total: int, target: ExcelDetailTarget) -> dict[str, Any]:
    started = time.perf_counter()
    logger.info(
        "[%s/%s] Excel detail source=%s row=%s",
        index,
        total,
        target.source_app,
        target.row_index,
    )
    result: dict[str, Any]
    try:
        open_url(resolve_short_url(target.url))
        result = scrape_record_content(target.row_index, source_app=target.source_app)
    except DeviceUnavailable:
        reset_device_session()
        raise
    except Exception as exc:
        logger.exception("Excel detail crawl failed row=%s", target.row_index)
        result = {
            "status": "error",
            "account_name": None,
            "read_count": 0,
            "comment_count": 0,
            "error": str(exc),
        }

    duration = round(time.perf_counter() - started, 2)
    account_name = result.get("account_name")
    status = result.get("status") or "error"
    account_ok = _valid_account(account_name)
    write_status = "success" if status == "success" and account_ok else status
    if status == "success" and not account_ok:
        write_status = "invalid_account"

    return {
        "row_index": target.row_index,
        "url": target.url,
        "source_app": target.source_app,
        "status": write_status,
        "raw_status": status,
        "account_name": account_name,
        "read_count": int(result.get("read_count") or 0),
        "comment_count": int(result.get("comment_count") or 0),
        "duration": duration,
        "capture_pages": result.get("capture_pages"),
        "ocr_attempted": result.get("ocr_attempted"),
        "error": result.get("error"),
    }


def _start_execution_for_target(
    input_path: Path,
    output_path: Path,
    sheet_name: str,
    target: ExcelDetailTarget,
) -> int | None:
    submission_id = upsert_excel_row_submission(
        path=str(input_path.resolve()),
        sheet_name=sheet_name,
        row_index=target.row_index,
        url=target.url,
        source_app=target.source_app,
        output_path=str(output_path.resolve()),
        task_type=DETAIL_CRAWL_TASK_TYPE,
        max_attempts=Config.DETAIL_MAX_RETRIES,
    )
    try:
        return start_task_execution(submission_id, worker_id="excel_detail")
    except ValueError as exc:
        logger.warning("Excel detail task submission not runnable row=%s: %s", target.row_index, exc)
        return None


def _finish_excel_execution(
    execution_id: int,
    *,
    item: dict[str, Any],
    writeback_status: str,
    writeback_locator: dict[str, Any],
) -> None:
    metrics = _execution_metrics(item)
    try:
        finish_task_execution(
            execution_id,
            status=item.get("status") or "error",
            account_name=item.get("account_name"),
            metrics=metrics,
            result=item,
            writeback_status=writeback_status,
            writeback_locator=writeback_locator,
            writeback_error=item.get("error"),
            error=item.get("error"),
        )
    except Exception as exc:
        logger.warning("failed to finish Excel task execution id=%s: %s", execution_id, exc)


def _update_excel_writeback(
    execution_id: int,
    *,
    writeback_status: str,
    writeback_locator: dict[str, Any],
) -> None:
    try:
        update_task_execution_writeback(
            execution_id,
            writeback_status=writeback_status,
            writeback_locator=writeback_locator,
        )
    except Exception as exc:
        logger.warning("failed to update Excel writeback id=%s: %s", execution_id, exc)


def _execution_metrics(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "row_index": item.get("row_index"),
        "read_count": int(item.get("read_count") or 0),
        "comment_count": int(item.get("comment_count") or 0),
        "duration": item.get("duration"),
        "capture_pages": item.get("capture_pages"),
        "ocr_attempted": item.get("ocr_attempted"),
    }


def _error_item(target: ExcelDetailTarget, error: str) -> dict[str, Any]:
    return {
        "row_index": target.row_index,
        "url": target.url,
        "source_app": target.source_app,
        "status": "error",
        "raw_status": "error",
        "account_name": None,
        "read_count": 0,
        "comment_count": 0,
        "duration": 0,
        "capture_pages": None,
        "ocr_attempted": None,
        "error": error,
    }


def _write_result(worksheet, item: dict[str, Any]) -> None:
    row_index = int(item["row_index"])
    status = item["status"]
    if status == "success":
        _set_cell(worksheet, row_index, Config.EXCEL_DETAIL_COL_ACCOUNT_NAME, item.get("account_name") or "")
        _set_cell(worksheet, row_index, Config.EXCEL_DETAIL_COL_READ_COUNT, item.get("read_count") or 0)
        _set_cell(worksheet, row_index, Config.EXCEL_DETAIL_COL_COMMENT_COUNT, item.get("comment_count") or 0)
    elif status in {"not_found", "deleted"}:
        _set_cell(worksheet, row_index, Config.EXCEL_DETAIL_COL_ACCOUNT_NAME, "N")

    _set_cell(worksheet, row_index, Config.EXCEL_DETAIL_COL_STATUS, status)
    _set_cell(worksheet, row_index, Config.EXCEL_DETAIL_COL_DURATION, item.get("duration"))
    _set_cell(worksheet, row_index, Config.EXCEL_DETAIL_COL_ERROR, item.get("error") or "")
    _set_cell(worksheet, row_index, Config.EXCEL_DETAIL_COL_SOURCE, item.get("source_app") or "")


def _summary(results: list[dict[str, Any]], total_duration: float) -> dict[str, Any]:
    by_source: dict[str, dict[str, Any]] = {}
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in results:
        grouped[item["source_app"]].append(item)
    for source_app, rows in sorted(grouped.items()):
        durations = [float(row.get("duration") or 0) for row in rows]
        by_source[source_app] = {
            "total": len(rows),
            "status": dict(Counter(row.get("status") for row in rows)),
            "avg_duration": round(sum(durations) / len(durations), 2) if durations else 0,
            "max_duration": max(durations) if durations else 0,
            "pages": dict(Counter(row.get("capture_pages") for row in rows)),
            "ocr": sum(1 for row in rows if row.get("ocr_attempted")),
        }
    return {
        "total": len(results),
        "duration": round(total_duration, 2),
        "by_source": by_source,
    }


def _has_existing_result(worksheet, row_index: int) -> bool:
    columns = (
        Config.EXCEL_DETAIL_COL_ACCOUNT_NAME,
        Config.EXCEL_DETAIL_COL_READ_COUNT,
        Config.EXCEL_DETAIL_COL_COMMENT_COUNT,
        Config.EXCEL_DETAIL_COL_STATUS,
    )
    return any(_cell_value(worksheet, row_index, column) for column in columns if column >= 0)


def _source_limit_reached(source_app: str, current_count: int) -> bool:
    limit = _SOURCE_LIMITS.get(source_app, 0)
    return bool(limit and current_count >= limit)


def _source_filter() -> set[str]:
    value = Config.EXCEL_DETAIL_SOURCE_FILTER.strip()
    if not value:
        return set()
    return {item.strip() for item in value.split(",") if item.strip()}


def _ensure_result_headers(worksheet) -> None:
    headers = {
        Config.EXCEL_DETAIL_COL_STATUS: "测试状态",
        Config.EXCEL_DETAIL_COL_DURATION: "耗时秒",
        Config.EXCEL_DETAIL_COL_ERROR: "错误",
        Config.EXCEL_DETAIL_COL_SOURCE: "链路类型",
    }
    for column, value in headers.items():
        if column >= 0 and not _cell_value(worksheet, 1, column):
            _set_cell(worksheet, 1, column, value)


def _should_save(index: int) -> bool:
    return index == 1 or index % max(Config.EXCEL_DETAIL_SAVE_EVERY, 1) == 0


def _valid_account(value: Any) -> bool:
    if not value:
        return False
    text = str(value).strip()
    if text in {"微信", "N"}:
        return False
    if any("\ue000" <= char <= "\uf8ff" for char in text):
        return False
    return len(text) <= 30


def _cell_value(worksheet, row_index: int, zero_based_col: int) -> str:
    if zero_based_col < 0:
        return ""
    value = worksheet.cell(row=row_index, column=zero_based_col + 1).value
    return "" if value is None else str(value).strip()


def _set_cell(worksheet, row_index: int, zero_based_col: int, value: Any) -> None:
    if zero_based_col < 0:
        return
    worksheet.cell(row=row_index, column=zero_based_col + 1).value = value


def _append_jsonl(path: Path, item: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(item, ensure_ascii=False) + "\n")
