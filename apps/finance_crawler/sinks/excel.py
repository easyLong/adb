"""Local Excel result sink adapter."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from apps.finance_crawler.config import Config
from apps.finance_crawler.domain.records import CrawlResult, WritebackResult


class ExcelSink:
    """Write crawl results back to a local XLSX sheet."""

    sink_type = "excel"

    def __init__(
        self,
        path: str | Path,
        *,
        save_as: str | Path | None = None,
        sheet_name: str | None = None,
        account_col: int | None = None,
        read_count_col: int | None = None,
        comment_count_col: int | None = None,
        status_col: int | None = None,
        screenshot_col: int | None = None,
    ) -> None:
        self.path = Path(path)
        self.save_as = Path(save_as) if save_as else self.path
        self.sheet_name = sheet_name
        self.account_col = Config.QQ_COL_ACCOUNT_NAME if account_col is None else account_col
        self.read_count_col = Config.QQ_COL_READ_COUNT if read_count_col is None else read_count_col
        self.comment_count_col = Config.QQ_COL_COMMENT_COUNT if comment_count_col is None else comment_count_col
        self.status_col = Config.QQ_COL_DETAIL_STATUS if status_col is None else status_col
        self.screenshot_col = Config.QQ_COL_SCREENSHOT if screenshot_col is None else screenshot_col

    def write_initial_check_results(self, rows: list[dict[str, Any]]) -> None:
        workbook, worksheet = self._open_sheet()
        try:
            for item in rows:
                row_index = int(item["row_index"])
                exists = bool(item.get("exists"))
                value = item.get("account_name") or ""
                if not exists:
                    value = "N"
                self._set_cell(worksheet, row_index, self.account_col, value)
            self._save(workbook)
        finally:
            workbook.close()

    def write_detail_results(self, rows: list[dict[str, Any]]) -> None:
        workbook, worksheet = self._open_sheet()
        try:
            for item in rows:
                row_index = int(item["row_index"])
                self._set_cell(worksheet, row_index, self.read_count_col, item.get("read_count"))
                self._set_cell(worksheet, row_index, self.comment_count_col, item.get("comment_count"))
                self._set_cell(worksheet, row_index, self.status_col, _detail_status(item))
                if self.screenshot_col >= 0 and item.get("screenshot_path"):
                    self._set_cell(worksheet, row_index, self.screenshot_col, item.get("screenshot_path"))
            self._save(workbook)
        finally:
            workbook.close()

    def write_results(self, results: list[CrawlResult]) -> list[WritebackResult]:
        rows: list[dict[str, Any]] = []
        output: list[WritebackResult] = []
        for result in results:
            row_index = result.metrics.get("row_index")
            if not row_index:
                output.append(
                    WritebackResult(
                        sink_type=self.sink_type,
                        status="skipped",
                        task_id=result.task_id,
                        error="missing row_index",
                    )
                )
                continue
            rows.append(
                {
                    "row_index": row_index,
                    "read_count": result.metrics.get("read_count"),
                    "comment_count": result.metrics.get("comment_count"),
                    "detail_status": result.status,
                    "screenshot_path": result.screenshot_path,
                }
            )
            output.append(
                WritebackResult(
                    sink_type=self.sink_type,
                    status="pending",
                    task_id=result.task_id,
                    locator={"path": str(self.save_as), "row_index": row_index},
                )
            )

        if rows:
            self.write_detail_results(rows)
            for item in output:
                if item.status == "pending":
                    item.status = "success"
        return output

    def _open_sheet(self):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Missing dependency: pip install openpyxl") from exc

        if not self.path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.path}")
        workbook = load_workbook(self.path)
        worksheet = workbook[self.sheet_name] if self.sheet_name else workbook.active
        return workbook, worksheet

    def _save(self, workbook) -> None:
        self.save_as.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self.save_as)

    @staticmethod
    def _set_cell(worksheet, row_index: int, zero_based_col: int, value: Any) -> None:
        if zero_based_col < 0:
            return
        worksheet.cell(row=row_index, column=zero_based_col + 1).value = value


def _detail_status(item: dict[str, Any]) -> Any:
    return item.get("detail_status")
