"""Local Excel source adapter."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from apps.finance_crawler.config import Config
from apps.finance_crawler.domain.records import SourceRecord
from apps.finance_crawler.utils import tabular_links


class ExcelSource:
    """Read candidate post links from a local XLSX sheet."""

    source_type = "excel"

    def __init__(
        self,
        path: str | Path,
        *,
        sheet_name: str | None = None,
        limit: int | None = None,
        post_time_col: int | None = None,
        url_col: int | None = None,
        sheet_title: str | None = None,
    ) -> None:
        self.path = Path(path)
        self.sheet_name = sheet_name
        self.limit = Config.FETCH_LIMIT if limit is None else limit
        self.post_time_col = Config.QQ_COL_POST_TIME if post_time_col is None else post_time_col
        self.url_col = Config.QQ_COL_URL if url_col is None else url_col
        self.sheet_title = sheet_title

    @property
    def source_name(self) -> str:
        sheet = self.sheet_name or "active"
        return f"{self.path.resolve()}:{sheet}"

    def fetch_candidates(self) -> list[dict[str, Any]]:
        rows, title = self._read_rows()
        candidates = tabular_links.eligible_candidates(
            rows,
            start_row=0,
            sheet_title=self.sheet_title or title,
            post_time_col=self.post_time_col,
            url_col=self.url_col,
        )
        if self.limit and self.limit > 0:
            candidates = candidates[: self.limit]
        return candidates

    def fetch_records(self) -> list[SourceRecord]:
        records: list[SourceRecord] = []
        for item in self.fetch_candidates():
            row_index = item["row_index"]
            records.append(
                SourceRecord(
                    record_id=f"{self.path.resolve()}:{self.sheet_name or 'active'}:{row_index}",
                    source_type=self.source_type,
                    source_name=self.source_name,
                    url=item["url"],
                    app_type=item.get("source_app"),
                    post_time=item.get("post_time"),
                    locator={
                        "path": str(self.path),
                        "sheet_name": self.sheet_name,
                        "row_index": row_index,
                    },
                    raw=_json_safe_item(item),
                )
            )
        return records

    def _read_rows(self) -> tuple[list[list[str]], str]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Missing dependency: pip install openpyxl") from exc

        if not self.path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.path}")

        workbook = load_workbook(self.path, data_only=True, read_only=True)
        try:
            worksheet = workbook[self.sheet_name] if self.sheet_name else workbook.active
            rows: list[list[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                rows.append(["" if value is None else str(value).strip() for value in row])
            return rows, str(worksheet.title or self.path.stem)
        finally:
            workbook.close()


def _json_safe_item(item: dict[str, Any]) -> dict[str, Any]:
    copied = dict(item)
    value = copied.get("post_time")
    if hasattr(value, "isoformat"):
        copied["post_time"] = value.isoformat(sep=" ")
    return copied
