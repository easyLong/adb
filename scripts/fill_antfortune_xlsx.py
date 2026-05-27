"""Fill local Ant Fortune XLSX rows using the existing device crawler."""

from __future__ import annotations

import argparse
import sys
import time
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from apps.finance_crawler.mobile.crawler import (
    check_record_exists_and_account,
    open_url,
    resolve_short_url,
    scrape_record_content,
)
from apps.finance_crawler.config import Config


ACCOUNT_HEADER = "发帖账号"
LINK_HEADER = "链接"
READ_HEADER = "阅读数"
COMMENT_HEADER = "评论数"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fill 发帖账号/阅读数/评论数 for a local Ant Fortune XLSX file."
    )
    parser.add_argument("input", help="Path to the source xlsx file.")
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name. Defaults to the active sheet.",
    )
    parser.add_argument(
        "--save-as",
        default=None,
        help="Output xlsx path. Defaults to <input>_filled.xlsx.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="Start processing from this row number.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Maximum number of rows to process. 0 means all eligible rows.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Process rows even when account/read/comment columns already have values.",
    )
    return parser.parse_args()


def find_header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is None:
            continue
        text = str(value).strip()
        if text:
            mapping[text] = col
    required = [ACCOUNT_HEADER, LINK_HEADER, READ_HEADER, COMMENT_HEADER]
    missing = [header for header in required if header not in mapping]
    if missing:
        raise RuntimeError(f"Missing required headers: {', '.join(missing)}")
    return mapping


def should_process(ws, row: int, header_map: dict[str, int], force: bool) -> bool:
    link = ws.cell(row, header_map[LINK_HEADER]).value
    if not link:
        return False
    if force:
        return True
    account = ws.cell(row, header_map[ACCOUNT_HEADER]).value
    read_count = ws.cell(row, header_map[READ_HEADER]).value
    comment_count = ws.cell(row, header_map[COMMENT_HEADER]).value
    return not any(value not in (None, "") for value in (account, read_count, comment_count))


def process_row(ws, row: int, header_map: dict[str, int]) -> dict[str, object]:
    url = str(ws.cell(row, header_map[LINK_HEADER]).value).strip()
    deep_link = resolve_short_url(url)
    open_url(deep_link)

    check_result = check_record_exists_and_account(row)
    if check_result["status"] == "not_found":
        ws.cell(row, header_map[ACCOUNT_HEADER]).value = "N"
        ws.cell(row, header_map[READ_HEADER]).value = 0
        ws.cell(row, header_map[COMMENT_HEADER]).value = 0
        return {
            "row": row,
            "url": url,
            "status": "not_found",
            "account_name": "N",
            "read_count": 0,
            "comment_count": 0,
        }
    if check_result["status"] == "error":
        return {
            "row": row,
            "url": url,
            "status": "error",
            "account_name": None,
            "read_count": None,
            "comment_count": None,
            "error": check_result.get("error"),
        }

    scrape_result = scrape_record_content(row, source_app="antfortune")
    account_name = check_result.get("account_name") or ""
    ws.cell(row, header_map[ACCOUNT_HEADER]).value = account_name
    ws.cell(row, header_map[READ_HEADER]).value = scrape_result.get("read_count") or 0
    ws.cell(row, header_map[COMMENT_HEADER]).value = scrape_result.get("comment_count") or 0
    return {
        "row": row,
        "url": url,
        "status": scrape_result["status"],
        "account_name": account_name,
        "read_count": scrape_result.get("read_count") or 0,
        "comment_count": scrape_result.get("comment_count") or 0,
        "error": scrape_result.get("error"),
    }


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = (
        Path(args.save_as)
        if args.save_as
        else input_path.with_name(f"{input_path.stem}_filled{input_path.suffix}")
    )

    workbook = load_workbook(input_path)
    worksheet = workbook[args.sheet] if args.sheet else workbook.active
    header_map = find_header_map(worksheet)

    processed = 0
    for row in range(args.start_row, worksheet.max_row + 1):
        if args.limit and processed >= args.limit:
            break
        if not should_process(worksheet, row, header_map, args.force):
            continue

        result = process_row(worksheet, row, header_map)
        processed += 1
        print(result)
        time.sleep(Config.POST_DELAY_MIN)

    workbook.save(output_path)
    print(f"saved={output_path}")
    print(f"processed={processed}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
